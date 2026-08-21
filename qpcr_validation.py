#!/usr/bin/env python3
"""qPCR template validation core — pure library, no UI dependency."""

import re, string
from pathlib import Path

# -- Constants -------------------------------------------------------

VALID_TASKS     = frozenset({"STND", "NTC", "UNKN"})
VALID_REPORTERS = {"SYBR", "FAM", "VIC", "JOE", "NED", "TAMRA", "TET", "ROX"}

MAX_SAMPLES_384      = 320
MAX_SAMPLES_96       = 70
MAX_DNA_SAMPLES_HARD = 192    # Hamilton PCR GUI hard cap for DNA rows
MAX_PRIMER_CUPLES    = 64     # unique FWD+REV name pairs

CHAR_ALLOWED = frozenset(string.ascii_letters + string.digits + " _-()")

# -- Shared helpers --------------------------------------------------

class Issue:
    """One validation finding."""
    OK   = "ok"
    WARN = "warn"
    FAIL = "fail"
    __slots__ = ("kind", "text")

    def __init__(self, kind, text):
        self.kind = kind
        self.text = text

    def __str__(self):
        return self.text

def is_blank(v):
    return v is None or (isinstance(v, str) and v.strip() == "")

def sanitize_name(name, ctx=""):
    """Require names to contain only ASCII alphanumerics, '_', '-' or parentheses."""
    issues = []
    invalid = sorted(set(c for c in name if c not in CHAR_ALLOWED))
    if invalid:
        cs = " / ".join(repr(c) for c in invalid)
        issues.append(Issue(Issue.FAIL,
                            f"{ctx}: '{name}' contains unsupported character(s): {cs}. "
                            "Use only letters, numbers, spaces, '_', '-' and parentheses."))
    return issues

def _load_sheet(filepath, expected):
    """Load *expected* sheet and return it with validation messages."""
    try:
        if filepath.endswith('.xls'):
            import xlrd as _xlrd
            wb = _xlrd.open_workbook(filepath)
            names = [s.name for s in wb.sheets()]
            sheet = next((s for s in wb.sheets() if s.name == expected), None)
        elif filepath.endswith(('.xlsx', '.xlsm')):
            import openpyxl as _opxl
            wb = _opxl.load_workbook(filepath, data_only=True)
            names = wb.sheetnames
            sheet = wb[expected] if expected in names else None
        else:
            return None, [Issue(Issue.FAIL,
                                f"'{filepath}' is not an Excel workbook type we can read. "
                                "Use an .xls, .xlsx, or .xlsm file.")]
    except Exception as e:
        return None, [Issue(Issue.FAIL,
                            f"Could not open '{Path(filepath).name}'. "
                            f"Check that the file is a valid Excel workbook. Details: {e}")]

    if sheet is None:
        return None, [Issue(Issue.FAIL,
                            f"{Path(filepath).name}: The '{expected}' sheet is missing. "
                            f"Available sheets: {names}")]
    return sheet, [Issue(Issue.OK, f"Sheet name '{expected}' is correct")]


def _read_all(sheet):
    """Pull every row as 4-element tuples."""
    if hasattr(sheet, 'nrows'):           # xlrd
        def cv(r, c):
            v = sheet.cell_value(r, c)
            if isinstance(v, float) and r > 0:
                return (str(int(v)) if v == int(v) else str(v)).strip()
            if isinstance(v, str):
                return str(v).strip()
            return v
        nr, nc = sheet.nrows, sheet.ncols
        return [tuple(cv(r, c) if c < nc else None for c in range(4))
                for r in range(nr)]
    else:                                 # openpyxl
        nr, nc = sheet.max_row, sheet.max_column
        return [tuple(sheet.cell(row=r + 1, column=c + 1).value
                      if c < nc else None for c in range(4))
                for r in range(nr)]


# -- DNA list validator (B.1) ----------------------------------------

def validate_dna_list(filepath):
    msgs = []
    seen = {}
    ntcs = []
    stds = []
    all_pos = []
    bn = Path(filepath).name

    if not filepath.endswith('.xls'):
        msgs.append(Issue(Issue.FAIL,
            f"{bn}: This file uses the wrong format for Hamilton. "
            "Save it as an .xls file before sending it to the robot."))

    sheet, sheet_msgs = _load_sheet(filepath, 'DNA')
    msgs.extend(sheet_msgs)
    if sheet is None:
        return msgs, 0

    rows = _read_all(sheet)
    expected_hdr = ('Position', 'Name', 'Task', 'Quantity')

    if not rows:
        msgs.append(Issue(Issue.FAIL,
                  f"{bn}: The workbook has no rows. Add the required header and at least one data row."))
        return msgs, 0

    hdr = tuple(str(h).strip() for h in rows[0] if h is not None)
    if len(hdr) != 4:
        msgs.append(Issue(Issue.FAIL,
                          f"{bn}: The header must have 4 columns: Position, Name, Task, Quantity. "
                          f"This file has {len(hdr)}."))
    elif hdr == expected_hdr:
        msgs.append(Issue(Issue.OK, f"{bn}: Header OK"))
    else:
        msgs.append(Issue(Issue.FAIL,
                          f"{bn}: The header is not correct. Expected {list(expected_hdr)}, "
                          f"but found {list(hdr)}."))

    pcap = 0
    for ri, row in enumerate(rows[1:], start=2):
        pos_v  = row[0] if len(row) > 0 else None
        name_v = row[1] if len(row) > 1 else None
        task_v = row[2] if len(row) > 2 else None
        qty_v  = row[3] if len(row) > 3 else None
        ctx = f"Row {ri}"

        if all(is_blank(v) for v in row):
            msgs.append(Issue(Issue.FAIL,
                f"{ctx}: This DNA row is completely empty. Delete the row or fill in all required values."))
            continue

        pcap += 1
        ns = str(name_v).strip() if not is_blank(name_v) else ""

        if is_blank(ns):
            msgs.append(Issue(Issue.FAIL,
                              f"{ctx}: The sample name is empty. Enter a unique sample name."))
            continue

        msgs.extend(sanitize_name(ns, ctx))

        ln = ns.lower()
        if ln in seen:
            msgs.append(Issue(Issue.FAIL,
                f"{ctx}: The sample name '{ns}' is used more than once. "
                f"The first occurrence is on row {seen[ln]}; every sample name must be unique."))
        else:
            seen[ln] = ri

        if pos_v and not is_blank(pos_v):
            try:
                all_pos.append(int(float(str(pos_v).strip())))
            except (ValueError, TypeError):
                pass

        ts = str(task_v).strip().upper() if not is_blank(task_v) else ""
        if not is_blank(ts):
            if ts not in VALID_TASKS:
                msgs.append(Issue(Issue.FAIL,
                    f"{ctx}: Task '{ts}' is not recognized. Enter UNKN, STND, or NTC."))
            if ts == "NTC":
                ntcs.append((ns, ri))
            elif ts == "STND":
                stds.append(ns)
                
        if is_blank(qty_v):
            msgs.append(Issue(Issue.FAIL,
                f"{ctx}: Quantity is empty. Enter a number; use 0 for NTC and UNKN rows."))
        else:
            qstr = str(qty_v).strip().replace(",", "")
            try:
                qv = float(qstr)
            except (ValueError, TypeError):
                qv = None
                msgs.append(Issue(Issue.FAIL,
                    f"{ctx}: Quantity '{qty_v}' is not a number. Enter a numeric value."))
            else:
                if ts in ("NTC", "UNKN") and qv != 0:
                    msgs.append(Issue(Issue.FAIL,
                        f"{ctx}: A {ts} row must have quantity 0, but this row has '{qty_v}'."))
                elif ts == "STND":
                    raw = str(qty_v).strip()
                    if "." in raw or "," in raw:
                        msgs.append(Issue(Issue.FAIL,
                            f"{ctx}: Standard quantity '{raw}' must be a whole number without commas or dots."))

    # --- Post-row checks --------------------------------------------------

    # Check for at least 1 DNA row, and enforce hard cap (192 samples) on total DNA rows
    total_dna_rows = pcap
    if total_dna_rows == 0:
        msgs.append(Issue(Issue.FAIL,
            f"{bn}: The DNA sheet has no data rows. Add at least one DNA sample row."))
        return msgs, 0
    elif total_dna_rows > MAX_DNA_SAMPLES_HARD:
        msgs.append(Issue(Issue.FAIL,
                f"{bn}: There are {total_dna_rows} DNA rows, but the maximum is "
                f"{MAX_DNA_SAMPLES_HARD}. Remove some rows."))
    else:
        msgs.append(Issue(Issue.OK,
            f"{bn}: Found {total_dna_rows} DNA rows "
            f"(maximum allowed: {MAX_DNA_SAMPLES_HARD})"))

    if not ntcs:
        msgs.append(Issue(Issue.OK,
            f"{bn}: No NTC controls found (optional)"))
    else:
        msgs.append(Issue(Issue.OK,
            f"{bn}: Found {len(ntcs)} NTC control(s)"))

    if not stds:
        msgs.append(Issue(Issue.OK,
            f"{bn}: No standard curve data found (optional)"))
    elif len(stds) < 6:
        msgs.append(Issue(Issue.WARN,
            f"{bn}: Found {len(stds)} standard(s). A standard curve should preferably "
            "have at least 6 standard samples."))
    else:
        msgs.append(Issue(Issue.OK,
            f"{bn}: Found {len(stds)} standard(s)"))

    if all_pos:
        mx = max(all_pos)
        sc = len(seen)
        is_384 = mx > 96
        fl   = "384-well" if is_384 else "96-well"
        lm   = MAX_SAMPLES_384 if is_384 else MAX_SAMPLES_96
        if sc > lm:
            msgs.append(Issue(Issue.FAIL,
                f"{bn}: There are {sc} unique samples, but a {fl} plate holds about {lm}. "
                "Reduce the number of samples or use a larger plate."))
        else:
            msgs.append(Issue(Issue.OK,
                f"{bn}: {sc} samples fits {fl} plate OK "
                f"(max_pos={mx}, limit={lm})"))

    # Check for sample rows without a Task
    has_incomplete = False
    for row in rows[1:]:
        nm = str(row[1]).strip() if not is_blank(row[1]) else ""
        ts = str(row[2]).strip().upper() if not is_blank(row[2]) else ""
        if nm and not ts:
            has_incomplete = True

    if has_incomplete:
        msgs.append(Issue(Issue.FAIL,
            f"{bn}: At least one row has a sample name but no Task. "
            "Enter UNKN, STND, or NTC in the Task column."))
    else:
        msgs.append(Issue(Issue.OK, f"{bn}: No incomplete rows"))

    return msgs, pcap


# -- PRIMER list validator (B.2) -------------------------------------

def validate_primer_list(filepath):
    msgs = []
    seen = {}
    all_pos = []
    bn = Path(filepath).name

    if not filepath.endswith('.xls'):
        msgs.append(Issue(Issue.FAIL,
            f"{bn}: This file uses the wrong format for Hamilton. "
            "Save it as an .xls file before sending it to the robot."))
    sheet, sheet_msgs = _load_sheet(filepath, 'PRIMER')
    msgs.extend(sheet_msgs)
    if sheet is None:
        return msgs, 0

    rows = _read_all(sheet)
    expected_hdr = ('Position', 'Name/Detector', 'Reporter', 'Quencher')

    if not rows:
        msgs.append(Issue(Issue.FAIL,
                  f"{bn}: The workbook has no rows. Add the required header and at least one primer row."))
        return msgs, 0

    hdr = tuple(str(h).strip() for h in rows[0] if h is not None)
    if len(hdr) != 4:
        msgs.append(Issue(Issue.FAIL,
                          f"{bn}: The header must have 4 columns: Position, Name/Detector, Reporter, Quencher. "
                          f"This file has {len(hdr)}."))
    elif hdr == expected_hdr:
        msgs.append(Issue(Issue.OK, f"{bn}: Primer header OK"))
    else:
        msgs.append(Issue(Issue.FAIL,
                          f"{bn}: The header is not correct. Expected {list(expected_hdr)}, "
                          f"but found {list(hdr)}."))

    pcap = 0
    for ri, row in enumerate(rows[1:], start=2):
        pos_v  = row[0] if len(row) > 0 else None
        nmv    = row[1] if len(row) > 1 else None
        repv   = row[2] if len(row) > 2 else None
        quv    = row[3] if len(row) > 3 else None
        ctx = f"Row {ri}"

        if all(is_blank(v) for v in row):
            msgs.append(Issue(Issue.FAIL,
                f"{ctx}: This PRIMER row is completely empty. Delete the row or fill in all required values."))
            continue

        name_str   = str(nmv).strip() if not is_blank(nmv) else ""
        report_str = str(repv).strip().upper() if not is_blank(repv) else ""
        quench_str = str(quv).strip() if not is_blank(quv) else ""

        if is_blank(name_str):
            msgs.append(Issue(Issue.FAIL,
                              f"{ctx}: The primer or assay name is empty. Enter a unique name."))
            continue

        msgs.extend(sanitize_name(name_str, ctx))
        ln = name_str.lower()
        if ln in seen:
            msgs.append(Issue(Issue.FAIL,
                f"{ctx}: The primer name '{name_str}' is used more than once. "
                f"The first occurrence is on row {seen[ln]}; every primer name must be unique."))
        else:
            seen[ln] = ri

        if pos_v and not is_blank(pos_v):
            try:
                all_pos.append(int(float(str(pos_v).strip())))
            except (ValueError, TypeError):
                pass

        if report_str:
            if report_str not in VALID_REPORTERS:
                msgs.append(Issue(Issue.FAIL,
                    f"{ctx}: Reporter '{report_str}' is not recognized. "
                    "Use one of the supported reporter names: SYBR, FAM, VIC, JOE, NED, TAMRA, TET, or ROX."))

        pcap += 1

    pc = len(seen)
    if all_pos:
        mx = max(all_pos)
        is_384 = mx > MAX_SAMPLES_96
        fl = "384-well" if is_384 else "96-well"
        lm = MAX_SAMPLES_384 if is_384 else MAX_SAMPLES_96
        if pc > lm:
            msgs.append(Issue(Issue.FAIL,
                f"{bn}: There are {pc} unique primers, but a {fl} plate holds about {lm}. "
                "Reduce the number of primers or use a larger plate."))
        else:
            msgs.append(Issue(Issue.OK,
                f"{bn}: {pc} primer(s) fits {fl} plate OK (max_pos={mx})"))

    # Hamilton PCR GUI hard cap for unique primer name-pairs
    if pc > MAX_PRIMER_CUPLES:
        msgs.append(Issue(Issue.FAIL,
            f"{bn}: There are {pc} unique primer name-pairs, but the maximum is "
            f"{MAX_PRIMER_CUPLES}. Remove some primer pairs."))
    elif pc:
        msgs.append(Issue(Issue.OK,
            f"{bn}: Found {pc} primer name-pairs "
            f"(maximum allowed: {MAX_PRIMER_CUPLES})"))

    return msgs, pcap


# -- File-type detection (A.) ----------------------------------------

def validate_sheet_selection(filepath):
    """Reject workbooks containing both DNA and PRIMER sheets."""
    if filepath.endswith('.csv'):
        return []

    try:
        if filepath.endswith('.xls'):
            import xlrd as _xlrd
            sheets = [s.name.lower() for s in _xlrd.open_workbook(filepath).sheets()]
        elif filepath.endswith(('.xlsx', '.xlsm')):
            import openpyxl as _opxl
            sheets = [s.lower() for s in _opxl.load_workbook(filepath, read_only=True).sheetnames]
        else:
            return []
    except Exception:
        return []

    if 'dna' in sheets and 'primer' in sheets:
        return [Issue(Issue.FAIL,
                      f"{Path(filepath).name}: This workbook contains both DNA and PRIMER sheets. "
                      "Use one workbook for DNA or one workbook for PRIMER, not both.")]
    return []

def detect_file_type(filepath):
    bn = Path(filepath).name.lower()
    if filepath.endswith('.csv'):
        return ['unknown']

    try:
        if filepath.endswith('.xls'):
            import xlrd as _xlrd
            wb = _xlrd.open_workbook(filepath)
            sheets = [s.name for s in wb.sheets()]
        elif filepath.endswith(('.xlsx', '.xlsm')):
            import openpyxl as _opxl
            wb = _opxl.load_workbook(filepath, data_only=True)
            sheets = wb.sheetnames
        else:
            return []

        types = []
        sheet_lower = {s.lower() for s in sheets}

        if 'dna' in sheet_lower:
            types.append('dna_list')
        if 'primer' in sheet_lower:
            types.append('primer_list')
        return types if types else ['unknown']
    except Exception:
        return ['read_error', 'unknown']