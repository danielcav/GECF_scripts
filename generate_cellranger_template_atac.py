#!/usr/bin/env python3
"""
generate_cellranger_template_atac.py

Usage:
    python3 generate_cellranger_template_atac.py file.xlsx

Reads an Excel sheet containing (at least) these columns:
    - "GECF fastq ID (avidxxxx)"
    - "Sample name" (trailing spaces in the header are tolerated)
and a cell somewhere in the sheet containing the run name (AVT0XXX).

It then asks which reference genome to use, and generates a
cellranger-atac_AVT0XXX.sh script pre-filled with one
"cellranger-atac count" block per sample.
"""

import argparse
import readline
import glob
import os
import re
import sys
import warnings

# Suppress harmless openpyxl warnings about unsupported Excel features
warnings.filterwarnings("ignore", message="Data Validation extension is not supported")
warnings.filterwarnings("ignore", message="Conditional Formatting extension is not supported")

try:
    import openpyxl
except ImportError:
    sys.exit(
        "Error: the python module 'openpyxl' is required.\n"
        "Install it with: pip3 install openpyxl"
    )


NGS_BASE_DIR_TEMPLATE = "/mnt/PTEGraw/AA/NGSruns/AVT/AV240401/{avt_run}"

HUMAN_GENOME_PATH = "/home/gecf/references/cellranger_references/current_human_arc_ATAC/refdata-cellranger-arc-GRCh38-2024-A"
MOUSE_GENOME_PATH = "/home/gecf/references/cellranger_references/current_mouse_arc_ATAC/refdata-cellranger-arc-GRCm39-2024-A"

AVT_PATTERN = re.compile(r"AVT0?\d{3,5}", re.IGNORECASE)


def _path_completer(text, state):
    matches = glob.glob(os.path.expanduser(text) + "*")
    matches = [m + "/" if os.path.isdir(m) else m for m in matches]
    return matches[state] if state < len(matches) else None


readline.set_completer_delims(" \t\n;")
readline.set_completer(_path_completer)
readline.parse_and_bind("tab: complete")


def norm(value):
    if value is None:
        return ""
    return str(value).strip().lower()


# ─── helpers that survive merged headers without openpyxl's broken detection ──

from collections import namedtuple
_Merged = namedtuple("_Merged", "min_row max_row min_col max_col")


def build_merged_cell_lookup(ws):
    """
    Returns a dict mapping (row, col) -> value for every cell that belongs
    to a merged range, using the value stored in the top-left cell of that
    range. This lets header rows that use merged cells (e.g. a label
    merged across two rows) still resolve correctly.
    """
    lookup = {}
    for merged_range in ws.merged_cells.ranges:
        min_row, min_col, max_row, max_col = (
            merged_range.min_row,
            merged_range.min_col,
            merged_range.max_row,
            merged_range.max_col,
        )
        top_left_value = ws.cell(row=min_row, column=min_col).value
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                lookup[(r, c)] = top_left_value
    return lookup


def get_cell_value(ws, merged_lookup, row_idx, col_idx):
    """
    Return the value seen at (row, col). If that position is part of a
    merged range whose top-left cell holds data, return that instead.
    Falls back gracefully if ws.cell itself throws.
    """
    try:
        # First check our lookup for merged ranges
        if (row_idx, col_idx) in merged_lookup:
            val = merged_lookup[(row_idx, col_idx)]
            if val is not None:
                return val
        return ws.cell(row=row_idx, column=col_idx).value
    except Exception:
        return None


# ─── main parsing logic ───────────────────────────────────────────────

def parse_xlsx(path):
    """
    Returns (sample_fastq_pairs, detected_avt_run) for the given Excel file.
    
    sample_fastq_pairs: list of (sample_name, fastq_id, avt_run) triples
    
    Three-level run-name detection:
      1) Per-row: look in the fastq_id column and adjacent columns on each data row
      2) Sheet-wide: full scan of all cells 
      3) Filename fallback: extract AVT pattern from the file name itself
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    print(f"  Sheets: {', '.join(wb.sheetnames)}")

    if "User_x1x_copy" in wb.sheetnames:
        ws = wb["User_x1x_copy"]
        print(f"  -> Using 'User_x1x_copy' sheet")
    else:
        ws = wb.active
        print(f"  -> Using active sheet '{ws.title}'")

    merged_lookup = build_merged_cell_lookup(ws)

    fastq_col = None
    sample_col = None
    header_row_idx = None

    # Scan header rows, combining each row with the row above it, so a
    # header label merged/split across two rows (e.g. "GECF fastq ID" on
    # one row and "(avidxxxx)" on the row above/below) is still detected.
    max_header_scan = min(15, ws.max_row)
    for row_idx in range(1, max_header_scan + 1):
        row_vals = [
            get_cell_value(ws, merged_lookup, row_idx, c)
            for c in range(1, ws.max_column + 1)
        ]
        prev_row_vals = (
            [
                get_cell_value(ws, merged_lookup, row_idx - 1, c)
                for c in range(1, ws.max_column + 1)
            ]
            if row_idx > 1
            else [""] * ws.max_column
        )
        combined = [
            f"{norm(prev)} {norm(cur)}".strip()
            for prev, cur in zip(prev_row_vals, row_vals)
        ]

        f_idx = None
        s_idx = None
        for i, v in enumerate(combined):
            if "fastq id" in v and "avid" in v:
                f_idx = i
            if "sample name" in v:
                s_idx = i
        if f_idx is not None and s_idx is not None:
            fastq_col = f_idx + 1
            sample_col = s_idx + 1
            header_row_idx = row_idx
            break

    if fastq_col is None or sample_col is None:
        sys.exit(
            "Error: could not find both 'GECF fastq ID (avidxxxx)' and "
            "'Sample name' columns in the first 10 rows."
        )

    # If the header cell(s) are merged across multiple rows, make sure we
    # start reading data *after* the full merged range, not right after
    # the row where we first detected the header text.
    data_start_row = header_row_idx + 1
    for merged_range in ws.merged_cells.ranges:
        if (
            merged_range.min_row <= header_row_idx <= merged_range.max_row
            and (
                merged_range.min_col <= fastq_col <= merged_range.max_col
                or merged_range.min_col <= sample_col <= merged_range.max_col
            )
        ):
            data_start_row = max(data_start_row, merged_range.max_row + 1)

    print(
        f"Found header on row {header_row_idx}: "
        f"fastq column {fastq_col}, sample column {sample_col}"
    )

    pairs = []
    blank_streak = 0
    for row_idx in range(data_start_row, ws.max_row + 1):
        fastq_val = get_cell_value(ws, merged_lookup, row_idx, fastq_col)
        sample_val = get_cell_value(ws, merged_lookup, row_idx, sample_col)

        fastq_val = str(fastq_val).strip() if fastq_val is not None else ""
        sample_val = str(sample_val).strip() if sample_val is not None else ""

        # Excel returns 0 for empty numeric-formatted cells — treat those
        # the same as genuinely blank.
        fastq_val = "" if fastq_val == "0" else fastq_val
        sample_val = "" if sample_val == "0" else sample_val

        # Stop immediately if we hit a "Run details" marker row (a common
        # section break some sheets use after the sample table)
        if "run details" in norm(fastq_val) or "run details" in norm(sample_val):
            break

        if fastq_val and sample_val:
            # Level 1 – per-row AVT detection: try fastq_id first, then other columns
            row_avt = None
            fq_match = AVT_PATTERN.search(fastq_val)
            if fq_match:
                row_avt = fq_match.group(0).upper()
            else:
                for c in range(1, ws.max_column + 1):
                    if c == fastq_col or c == sample_col:
                        continue
                    val = get_cell_value(ws, merged_lookup, row_idx, c)
                    if val is not None:
                        cm = AVT_PATTERN.search(str(val))
                        if cm:
                            row_avt = cm.group(0).upper()
                            break

            pairs.append((sample_val, fastq_val, row_avt))
            blank_streak = 0
        elif not fastq_val and not sample_val:
            # Fully empty row: could be a stray blank row, or the real end
            # of the table. Tolerate isolated blank rows, but stop once we
            # see two blank rows in a row (after we've already collected
            # at least one pair).
            blank_streak += 1
            if pairs and blank_streak >= 2:
                break
        else:
            # One of the two cells is empty but not both: likely a
            # malformed/partial row. Skip it but keep going.
            print(
                f"Warning: row {row_idx} has only one of "
                f"fastq ID / sample name filled in; skipping."
            )
            blank_streak = 0

    # Level 2 – search whole sheet for an AVT run name
    avt_found = None
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            m = AVT_PATTERN.search(str(cell.value))
            if m:
                avt_found = m.group(0).upper()
                break
        if avt_found:
            break

    # Level 3 – fallback: extract AVT from the file name itself
    if not avt_found:
        fname_match = AVT_PATTERN.search(os.path.basename(path))
        if fname_match:
            avt_found = fname_match.group(0).upper()
            print(f"  -> Detected run '{avt_found}' from file name")

    return pairs, avt_found


def print_pairs_table(pairs):
    """
    Print a simple aligned table of (sample_name, fastq_id, avt_run) triples,
    with the fastq ID column first.
    """
    if not pairs:
        return

    fastq_header = "Fastq ID"
    sample_header = "Sample name"
    avt_header = "AVT run"

    fastq_width = max(len(fastq_header), max(len(f) for _, f, _ in pairs))
    sample_width = max(len(sample_header), max(len(s) for s, _, _ in pairs))
    avt_width = max(len(avt_header), max(len(a or "(none)") for _, _, a in pairs))

    sep = f"+{'-' * (fastq_width + 2)}+{'-' * (sample_width + 2)}+{'-' * (avt_width + 2)}+"

    print(sep)
    print(f"| {fastq_header:<{fastq_width}} | {sample_header:<{sample_width}} | {avt_header:<{avt_width}} |")
    print(sep)
    for sample_name, fastq_id, avt in pairs:
        print(f"| {fastq_id:<{fastq_width}} | {sample_name:<{sample_width}} | {(avt or '(none)'):<{avt_width}} |")
    print(sep)


# ─── user prompts (suppressed when NON_INTERACTIVE) ────────────────────

# Global flag for non-interactive execution (CI/testing)
NON_INTERACTIVE = False


def ask_avt_run(detected_avt):
    if NON_INTERACTIVE:
        return detected_avt or "AVT0000"
    if detected_avt:
        while True:
            print(f"Detected run name: {detected_avt}")
            answer = input("Use this run name? [Y/n]: ").strip()
            if answer[0].lower() == "y":
                return detected_avt
            elif answer[0].lower() == "n":
                break
            else:
                print("Answer can only be Y or n. Please try again.")
                continue

    while True:
        avt_run = input("Enter AVT run name (e.g. AVT0123): ").strip()
        if avt_run:
            return avt_run.upper()
        else:
            print("Run name cannot be empty. Please try again.\n")


def ask_reference_genome():
    genome_choices = [
        ("Human ARC ATAC", HUMAN_GENOME_PATH),
        ("Mouse ARC ATAC", MOUSE_GENOME_PATH),
    ]

    if NON_INTERACTIVE:
        return HUMAN_GENOME_PATH

    while True:
        print("\nAvailable reference genomes:")
        for i, (label, path) in enumerate(genome_choices, start=1):
            print(f"  {i}. {label} — {path}")
        print("  Custom: type a full path")
        answer = input("Choice [1]: ").strip() or "1"

        if answer == "1":
            return HUMAN_GENOME_PATH
        elif answer == "2":
            return MOUSE_GENOME_PATH
        else:
            custom_path = os.path.expanduser(answer)
            if os.path.isdir(custom_path):
                return custom_path
            else:
                print(f"Path '{custom_path}' does not exist. Try again.\n")


def build_script(reference_genome, pairs):
    """Generate the cellranger-atac shell script content."""

    # Derive a main AVT for output filenames (most common one, or MIXED)
    from collections import Counter
    avt_counts = Counter(avt for _, _, avt in pairs if avt is not None)
    main_avt = avt_counts.most_common(1)[0][0] if avt_counts else "MIXED_runs"

    # Build per-sample fastq path variables with timestamps and custom output dirs
    NGS_BASE_DIR_TEMPLATE = "/mnt/PTEGraw/AA/NGSruns/AVT/AV240401/{avt_run}"

    lines.append(f"# Generated by generate_cellranger_template_atac.py")
    lines.append("# Review fastq paths, fill in Server_folder and NGSruns_folder before running.")
    lines.append("")

    # Build per-fastq variables. Each uses the AVT embedded in its path string.
    for idx, (sample_name, fastq_id, avt) in enumerate(pairs, start=1):
        run_part = avt or main_avt
        base_dir = NGS_BASE_DIR_TEMPLATE.format(avt_run=run_part)
        lines.append(
            f"# --- {idx}. {sample_name} ({fastq_id}, AVT: {run_part}) ---"
        )
        # We will fill in the full path interactively or after review
        avt_path = base_dir  # Will need manual fine-tuning for exact subfolder
        lines.append(
            f"# fastq_path_{idx}={avt_path}/{fastq_id}"
        )

    return "\n".join(lines) + "\n"


# Global list for building output script
lines = []


def find_fastq_path(avt_run, fastq_id):
    """
    Search recursively under the run's base directory for a directory
    whose name is exactly `fastq_id` (it can live under any subfolder
    structure). Returns the full path if found, or None if not found
    (e.g. base dir doesn't exist on this machine, or no matching folder).
    """
    base_dir = NGS_BASE_DIR_TEMPLATE.format(avt_run=avt_run)

    if not os.path.isdir(base_dir):
        return None

    for root, dirnames, _ in os.walk(base_dir):
        for d in dirnames:
            if d == fastq_id:
                return os.path.join(root, d)

    return None


def build_script_full(reference_genome, pairs):
    """Build complete cellranger-atac shell script."""
    global lines
    lines = []

    # Derive a main AVT for output filenames (most common one, or MIXED)
    from collections import Counter
    avt_counts = Counter(avt for _, _, avt in pairs if avt is not None)
    main_avt = avt_counts.most_common(1)[0][0] if avt_counts else "MIXED_runs"

    lines.append("#!/bin/bash")
    lines.append("TIMESTAMP=$(date +%Y%m%d_%H%M%S)")
    lines.append("")

    # Group fastq IDs per sample name so each unique sample gets ONE count call
    from collections import defaultdict, Counter as Ctr

    samples_to_fastqs = defaultdict(list)
    for sn, fq, avt in pairs:
        samples_to_fastqs[sn].append((fq, avt))

    unique_samples = sorted(samples_to_fastqs.keys())

    # Probe real filesystem for each fastq dir and build path variables
    any_missing = False
    for sn in unique_samples:
        fastqs = samples_to_fastqs[sn]
        lines.append("# --- " + sn + " (" + str(len(fastqs)) + " FASTQ(s)) ---")
        for fq, avt in fastqs:
            run_part = avt or main_avt
            base_dir = NGS_BASE_DIR_TEMPLATE.format(avt_run=run_part)

            resolved_path = find_fastq_path(run_part, fq)
            if resolved_path:
                lines.append("path_" + sn + "_" + fq + "=" + resolved_path)
            else:
                lines.append("# path_" + sn + "_" + fq + "=" + base_dir + "/FIXME_LOCATE_" + fq)
                if os.path.isdir(base_dir):
                    print(
                        "  Warning: no folder named '" + fq + "' found under "
                        + base_dir + " - path_" + sn + "_" + fq + " needs to be set manually."
                    )
                else:
                    any_missing = True
        lines.append("")

    if any_missing:
        print("  Note: some base directories are not on this machine —")
        print("  all fastq_path variables under those runs need to be set manually.")
    lines.append("")
    lines.append("# Extra cellranger-atac options (leave empty if none)")
    lines.append("# Examples: --chemistry=ARC-v1  --no-bam  --min-atac-count=500")
    lines.append("CUSTOM_OPTIONS=")
    lines.append("")

    # Generate working folder name with timestamp (expanded in Python, not bash)
    server_folder_name = f"{main_avt}_CRatac_${{TIMESTAMP}}"
    
    lines.append("")
    lines.append("# Working directory: cellranger-atac output and post-processing live here")
    lines.append(f"Server_folder=$PWD/{server_folder_name}")
    lines.append('mkdir -vp "$Server_folder"')
    lines.append("")
    lines.append("# Final archive destination (set this to the permanent NGS runs path)")
    example_base_dir = NGS_BASE_DIR_TEMPLATE.format(avt_run=main_avt or "AVTXXXX")
    lines.append(f"# example: {example_base_dir}/CRatac_${{TIMESTAMP}}")
    lines.append("NGSruns_folder=")
    lines.append("")
    lines.append("# Check that the reference genome directory exists")
    lines.append('if [[ ! -d "${reference_genome_used}" ]]; then')
    lines.append("    echo \"Reference genome '${reference_genome_used}' does not exist. Aborting.\" >&2")
    lines.append("    exit 1")
    lines.append("fi")
    lines.append("")
    lines.append("")

    for idx, sn in enumerate(unique_samples, start=1):
        fastqs = samples_to_fastqs[sn]
        numbered_dir = f"{idx:02d}_{sn}_atac"
        fq_list_str = ",".join(fq for fq, _ in fastqs)

        lines.append("# --- Sample " + str(idx) + ": " + sn + " (" + str(len(fastqs)) + " FASTQ(s)) ---")
        lines.append("cellranger-atac count --id=" + sn)
        lines.append("                       --description=\"" + fq_list_str + "\" \\")
        lines.append("                       --reference=${reference_genome_used} \\")

        # Build space-separated path list for all FASTQ dirs of this sample
        fq_paths = " ".join("${path_" + sn + "_" + fq + "}" for fq, _ in fastqs)
        lines.append("                       --fastqs=" + fq_paths + r" \\")
        lines.append("                       --output=${Server_folder}/" + numbered_dir + r" \\")
        lines.append("                       ${CUSTOM_OPTIONS}")
        lines.append("")

        lines.append("# Clean up SC_ATAC_COUNTER_CS (non-useful)")
        lines.append(
            "rm -r ${Server_folder}/" + numbered_dir + "/SC_ATAC_COUNTER_CS 2>/dev/null || true"
        )
        lines.append("")

    post_avt = main_avt or "MIXED_runs"
    # Post-processing in Server_folder; rsync everything to NGSruns_folder at the end
    lines.append("")
    lines.append("# ------------------------------------ POST-PROCESSING ------------------------------------ #")
    lines.append("")
    lines.append("# Collect per-sample web summary HTMLs into one folder for easy browsing")
    lines.append(f"mkdir -p ${{Server_folder}}/{post_avt}_Summaries/web_summaries")
    lines.append("")
    for idx, sn in enumerate(unique_samples, start=1):
        ndir = f'{idx:02d}_{sn}_atac/outs'
        lines.append(
            f"cp -n ${{Server_folder}}/{ndir}/web_summary.html "
            f"${{Server_folder}}/{post_avt}_Summaries/web_summaries/{sn}_WebSummary.html 2>/dev/null || true"
        )
    lines.append("")

    # Merge per-sample summary.csv (run-level QC metrics) across all samples into one file
    lines.append(
        "# Merge run-level QC metrics (summary.csv) from each sample into a single consolidated table"
    )
    lines.append(
        f'merged_csv="${{Server_folder}}/{post_avt}_Summaries/all_summary.csv"'
    )
    lines.append("header_csv_set=false")
    for idx, sn in enumerate(unique_samples, start=1):
        ndir = f'{idx:02d}_{sn}_atac/outs'
        src = f'${{Server_folder}}/{ndir}/summary.csv'
        lines.append(
            'if [ -f "' + src + '" ]; then '
            'if [ "$header_csv_set" = false ]; then '
            'cp "' + src + '" "$merged_csv"; '
            'header_csv_set=true; else '
            'tail -n +2 "' + src + '" >> "$merged_csv"; fi; '
            'else echo "WARNING: ' + src + ' not found" >&2; fi'
        )
    lines.append("")

    # Rsync everything from Server_folder to NGSruns_folder in one shot
    lines.append("")
    lines.append("# ------------------------------------ DEPLOY TO NGS RUNS FOLDER ------------------------------------ #")
    lines.append("")
    lines.append("# Rsync all results (sample dirs + Summaries folder) to the permanent location in one command")
    lines.append("rsync -a --progress ${Server_folder}/ ${NGSruns_folder}")
    lines.append("")
    return "\n".join(lines) + "\n"


# ─── main entry point ─────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description=(
            "Parses one or more Excel sheets and generates a cellranger-atac_AVT0XXX.sh "
            "script with one \"cellranger-atac count\" block per sample."
        ),
        epilog="Example:\n  python3 %(prog)s samples.xlsx\n  python3 %(prog)s run1.xlsx run2.xlsx run3.xlsx",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "xlsx_files",
        nargs="+",
        help=(
            "One or more Excel files containing at least the columns "
            "'GECF fastq ID (avidxxxx)' and 'Sample name'. "
            "A cell containing the run name (e.g. AVT0226) is also "
            "searched for automatically."
        ),
    )
    parser.add_argument(
        "--non-interactive", "-n",
        action="store_true",
        default=False,
        help="Skip all prompts, use detected values and human genome as defaults.",
    )
    args = parser.parse_args()

    global NON_INTERACTIVE
    NON_INTERACTIVE = args.non_interactive
    xlsx_files = args.xlsx_files

    all_pairs = []

    for xlsx_file in xlsx_files:
        if not os.path.isfile(xlsx_file):
            sys.exit(f"Error: file '{xlsx_file}' does not exist.")
        if not xlsx_file.lower().endswith(".xlsx"):
            sys.exit(f"Error: '{xlsx_file}' does not look like an .xlsx file.")

        pairs, detected_avt = parse_xlsx(xlsx_file)
        if not pairs:
            print(f"Warning: no sample/fastq ID pairs found in '{xlsx_file}' — skipping.")
            continue

        # Backfill any pair whose per-row AVT is None with the sheet-level detection
        all_pairs.extend(
            (s, f, a if a is not None else detected_avt)
            for s, f, a in pairs
        )

    if not all_pairs:
        sys.exit("Error: no sample/fastq ID pairs could be parsed from the provided file(s).")

    print()
    print(f"Total: {len(all_pairs)} sample(s) from {len(xlsx_files)} Excel file(s):")
    print_pairs_table(all_pairs)
    print()

    # Assign a run name to any pairs that have None
    needs_avt = [i for i, (_, _, avt) in enumerate(all_pairs) if avt is None]
    if needs_avt:
        known_avts = list(dict.fromkeys(avt for _, _, avt in all_pairs if avt is not None))
        detected_avt_hint = known_avts[0] if known_avts else None

        print(f"{len(needs_avt)} sample(s) have no AVT run assigned.")
        if detected_avt_hint:
            avt_run = ask_avt_run(detected_avt_hint)
        else:
            avt_run = ask_avt_run(None)

        if not avt_run:
            sys.exit("Error: AVT run name cannot be empty.")

        for i in needs_avt:
            all_pairs[i] = (all_pairs[i][0], all_pairs[i][1], avt_run)

    # Derive a representative AVT for the output filename: pick the most common one
    from collections import Counter
    avt_counts = Counter(avt for _, _, avt in all_pairs if avt is not None)
    main_avt = avt_counts.most_common(1)[0][0] if avt_counts else "MIXED_runs"

    reference_genome = ask_reference_genome()

    script_content = build_script_full(reference_genome, all_pairs)

    out_script = f"cellranger-atac_{main_avt}.sh"
    with open(out_script, "w") as f:
        f.write(script_content)
    os.chmod(out_script, 0o755)

    print(f"Generated: {out_script}")
    print("Review it, fill in Server_folder / NGSruns_folder, and adjust fastq paths if needed.")


if __name__ == "__main__":
    main()
